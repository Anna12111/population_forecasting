"""Первичная загрузка и подготовка данных WPP для приложения"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"
CSV_PATH = DATA_DIR / "population_data.csv"
SOURCE_META_PATH = DATA_DIR / "wpp_source.json"

PRIMARY_URL_TEMPLATE = (
    "https://population.un.org/wpp/assets/Excel%20Files/1_Indicator%20(Standard)/"
    "EXCEL_FILES/1_General/WPP{year}_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"
)
LEGACY_URL_TEMPLATE = (
    "https://population.un.org/wpp/Download/Files/1_Indicator%20(Standard)/"
    "EXCEL_FILES/1_General/WPP{year}_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"
)

USER_AGENT = "population-forecasting-init/1.0"


def configure_output_encoding() -> None:
    """Принудительно включает UTF-8 для служебных сообщений модуля запуска"""
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


configure_output_encoding()


def emit(message: str) -> None:
    """Печатает служебное сообщение так, чтобы его мог прочитать графический модуль запуска"""
    print(message, flush=True)


def emit_status(text: str) -> None:
    """Отправляет текстовый статус текущего этапа инициализации"""
    emit(f"STATUS|{text}")


def emit_progress(value: int, maximum: int = 100) -> None:
    """Отправляет процент выполнения текущего этапа"""
    value = max(0, min(int(value), int(maximum)))
    emit(f"PROGRESS|{value}|{maximum}")


def revision_years() -> list[int]:
    """Формирует список возможных ревизий WPP от текущего года к более старым"""
    current_year = datetime.now().year
    years = list(range(current_year, 2023, -1))
    if 2022 not in years:
        years.append(2022)
    return years


def candidate_urls() -> Iterable[tuple[int, str]]:
    """Возвращает набор официальных URL, которые проверяются при поиске актуального WPP"""
    for year in revision_years():
        for template in (PRIMARY_URL_TEMPLATE, LEGACY_URL_TEMPLATE):
            yield year, template.format(year=year)


def existing_wpp_files() -> list[Path]:
    """Ищет уже скачанные Excel-файлы WPP в каталоге data"""
    if not DATA_DIR.exists():
        return []
    files = list(DATA_DIR.glob("WPP*_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"))
    return sorted(files, key=extract_revision_year, reverse=True)


def extract_revision_year(path: Path) -> int:
    """Извлекает год ревизии WPP из имени файла"""
    match = re.search(r"WPP(\d{4})", path.name)
    return int(match.group(1)) if match else 0


def test_url(url: str, timeout: int = 15) -> bool:
    """Проверяет доступность файла на сервере без скачивания всего архива"""
    request = Request(url, headers={"User-Agent": USER_AGENT, "Range": "bytes=0-0"})
    try:
        with urlopen(request, timeout=timeout) as response:
            return response.status in (200, 206)
    except HTTPError as exc:
        return exc.code == 206
    except (URLError, TimeoutError, OSError):
        return False


def resolve_latest_download() -> tuple[int, str]:
    """Подбирает наиболее новую доступную ссылку на полный Excel-файл WPP"""
    emit_status("Поиск актуального файла WPP на сайте ООН")
    for index, (year, url) in enumerate(candidate_urls(), start=1):
        emit_progress(min(95, index * 7), 100)
        if test_url(url):
            emit_status(f"Найден WPP{year}")
            return year, url
    reserve_year = 2024
    reserve_url = PRIMARY_URL_TEMPLATE.format(year=reserve_year)
    emit_status("Не удалось проверить новые версии, используется резервная ссылка WPP2024")
    return reserve_year, reserve_url


def download_file(url: str, destination: Path) -> None:
    """Скачивает Excel-файл WPP и показывает прогресс по объёму загруженных данных"""
    temp_path = destination.with_suffix(destination.suffix + ".part")
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=60) as response:
        total = int(response.headers.get("Content-Length") or 0)
        downloaded = 0
        chunk_size = 1024 * 1024
        with temp_path.open("wb") as file:
            while True:
                chunk = response.read(chunk_size)
                if not chunk:
                    break
                file.write(chunk)
                downloaded += len(chunk)
                if total:
                    emit_progress(int(downloaded / total * 100), 100)
                else:
                    emit_progress((downloaded // chunk_size) % 100, 100)
    temp_path.replace(destination)


def ensure_wpp_file() -> Path:
    """Проверяет наличие WPP-файла, а при отсутствии скачивает его с сайта ООН"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    files = existing_wpp_files()
    if files:
        emit_status(f"Найден локальный файл {files[0].name}")
        emit_progress(100, 100)
        return files[0]

    year, url = resolve_latest_download()
    destination = DATA_DIR / f"WPP{year}_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"
    emit_status(f"Скачивание {destination.name}")
    emit_progress(0, 100)
    download_file(url, destination)
    SOURCE_META_PATH.write_text(
        json.dumps(
            {
                "revision_year": year,
                "url": url,
                "downloaded_at": datetime.now().isoformat(timespec="seconds"),
                "filename": destination.name,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    emit_status(f"Файл {destination.name} загружен")
    emit_progress(100, 100)
    return destination


def find_header_row(worksheet: Any, max_rows: int = 40) -> int:
    """Находит строку заголовков в листе WPP по ключевым колонкам Variant и Year"""
    for row_index in range(1, max_rows + 1):
        values = [cell.value for cell in worksheet[row_index]]
        if "Variant" in values and "Year" in values:
            return row_index
    raise RuntimeError("Не удалось найти строку заголовков на листе Estimates.")


def cell_to_value(value: Any) -> Any:
    """Готовит значение ячейки Excel к записи в CSV"""
    if value is None:
        return ""
    if isinstance(value, float):
        return value
    return value


def build_population_csv(wpp_path: Path) -> None:
    """Создаёт population_data.csv из листа Estimates без служебной шапки WPP"""
    needs_rebuild = not CSV_PATH.exists() or CSV_PATH.stat().st_mtime < wpp_path.stat().st_mtime
    if not needs_rebuild:
        emit_status("Локальный CSV уже подготовлен")
        emit_progress(100, 100)
        return

    emit_status("Подготовка population_data.csv из листа Estimates")
    emit_progress(0, 100)

    workbook = load_workbook(wpp_path, read_only=True, data_only=True)
    try:
        if "Estimates" not in workbook.sheetnames:
            raise RuntimeError("В файле WPP не найден лист Estimates.")
        worksheet = workbook["Estimates"]
        header_row = find_header_row(worksheet)
        header = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        header = [str(value).strip() if value is not None else "" for value in header]

        has_index = bool(header and header[0] == "Index")
        output_header = header if has_index else ["Index"] + header

        with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(output_header)
            written = 0
            total_rows = max(1, worksheet.max_row - header_row)
            for position, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=1):
                row_values = [cell_to_value(value) for value in row]
                if not any(value != "" for value in row_values):
                    continue
                if "Year" in header:
                    year_value = row_values[header.index("Year")]
                    if year_value in (None, ""):
                        continue
                written += 1
                writer.writerow(row_values if has_index else [written] + row_values)
                if position % 250 == 0:
                    emit_progress(int(position / total_rows * 100), 100)
    finally:
        workbook.close()

    emit_status(f"CSV подготовлен: {CSV_PATH.name}")
    emit_progress(100, 100)


def main() -> None:
    """Запускает проверку и подготовку локальных данных для приложения"""
    start = time.time()
    try:
        wpp_path = ensure_wpp_file()
        build_population_csv(wpp_path)
        emit_status(f"Данные готовы ({time.time() - start:.1f} сек.)")
        emit("DONE")
    except Exception as exc:
        emit(f"ERROR|{exc}")
        raise


if __name__ == "__main__":
    main()
