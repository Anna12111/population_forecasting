"""Загрузка прогнозных сценариев ООН WPP для внешнего сравнения результатов модели"""
import hashlib
import os
import pickle
import re
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

UN_WPP_FILENAME = "WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../data"))

UN_FORECAST_VARIANT_DESCRIPTIONS: Dict[str, str] = {
    "Medium variant": "Средний сценарий рождаемости ООН; базовый вариант для большинства сравнений.",
    "High variant": "Высокий сценарий рождаемости ООН.",
    "Low variant": "Низкий сценарий рождаемости ООН.",
    "Constant-fertility": "Рождаемость фиксируется на уровне базового периода.",
    "Instant-replacement": "Рождаемость мгновенно переводится к уровню простого воспроизводства.",
    "Instant-replacement zero migr": "Уровень простого воспроизводства + нулевая миграция.",
    "Momentum": "Инерционный сценарий: мгновенное замещение, постоянная смертность и нулевая миграция.",
    "Zero-migration": "Сценарий нулевой миграции ООН.",
    "No change": "Без изменений: постоянные рождаемость и смертность.",
    "No fertility below age 18": "Сценарий без рождаемости у женщин младше 18 лет.",
    "Accelerated ABR decline": "Сценарий ускоренного снижения adolescent birth rate.",
    "Accelarated ABR decline recup": "Сценарий ускоренного снижения ABR с последующей компенсацией.",
}

_COLUMN_MAP = {
    "Variant": "variant",
    "Region, subregion, country or area *": "country",
    "Year": "year",
    "Total Population, as of 1 January (thousands)": "population",
    "Births (thousands)": "births",
    "Total Deaths (thousands)": "deaths",
    "Net Number of Migrants (thousands)": "net_migration",
}
_NUMERIC_COLUMNS = ["population", "births", "deaths", "net_migration"]
_REQUIRED_HEADER_NAMES = set(_COLUMN_MAP.keys())
_NON_FORECAST_SHEETS = {"Estimates", "Notes", "Cover", "Contents", "Index", "Metadata", "Info"}


def get_un_forecast_variants() -> List[str]:
    """
    Возвращает список доступных прогнозных сценариев WPP

    При наличии Excel-файла список формируется по его листам
    В список включаются только те листы, где найдены необходимые демографические колонки
    Если файл данных ещё не подготовлен, используется стандартный набор сценариев
    """
    try:
        return _get_available_forecast_sheets(get_wpp_path())
    except Exception:
        return list(UN_FORECAST_VARIANT_DESCRIPTIONS.keys())


def get_un_forecast_description(variant: Optional[str]) -> str:
    """Возвращает короткое пояснение сценария ООН для подсказки в интерфейсе"""
    if not variant:
        return ""
    return UN_FORECAST_VARIANT_DESCRIPTIONS.get(
        variant,
        "Прогнозный сценарий ООН WPP. Описание сформировано автоматически, так как сценарий найден в файле данных.",
    )


def get_effective_un_forecast_variant(variant: Optional[str]) -> Optional[str]:
    """
    Возвращает сценарий WPP, доступный в текущем файле данных

    Если переданный сценарий найден в Excel-файле, функция возвращает его без изменений
    Если сценарий отсутствует, выбирается Medium variant или первый доступный прогнозный лист
    """
    variants = get_un_forecast_variants()
    if not variants:
        return variant
    if variant in variants:
        return variant
    if "Medium variant" in variants:
        return "Medium variant"
    return variants[0]


def _find_wpp_files(directory: str) -> List[str]:
    """Ищет файлы WPP в каталоге данных и сортирует их от более новой версии к старой"""
    if not os.path.isdir(directory):
        return []
    files = []
    for filename in os.listdir(directory):
        if filename.startswith("WPP") and filename.endswith("_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"):
            files.append(os.path.join(directory, filename))
    return sorted(files, key=lambda item: _extract_revision_year(item), reverse=True)


def _extract_revision_year(path: str) -> int:
    """Извлекает год ревизии WPP из имени файла, если он указан"""
    match = re.search(r"WPP(\d{4})", os.path.basename(path))
    return int(match.group(1)) if match else 0


def get_wpp_path() -> str:
    """Находит актуальный файл WPP в стандартных каталогах проекта или по переменной окружения"""
    env_path = os.getenv("WPP_FORECAST_FILE") or os.getenv("WPP2024_FORECAST_FILE")
    candidates: List[str] = []
    if env_path:
        candidates.append(env_path)

    module_dir = os.path.dirname(__file__)
    candidates.extend(_find_wpp_files(os.path.join(module_dir, "../../data")))
    candidates.extend(_find_wpp_files(os.path.join(os.getcwd(), "data")))
    candidates.extend(
        [
            os.path.join(module_dir, "../../data", UN_WPP_FILENAME),
            os.path.join(os.getcwd(), "data", UN_WPP_FILENAME),
            os.path.join(os.getcwd(), UN_WPP_FILENAME),
            os.path.join("/mnt/data", UN_WPP_FILENAME),
        ]
    )

    seen = set()
    for path in candidates:
        abs_path = os.path.abspath(path)
        if abs_path in seen:
            continue
        seen.add(abs_path)
        if os.path.exists(abs_path):
            return abs_path

    raise FileNotFoundError(
        "Файл прогноза ООН не найден. Запустите launch.pyw для первичной загрузки WPP "
        "или задайте переменную окружения WPP_FORECAST_FILE."
    )


def get_wpp2024_path() -> str:
    """Возвращает путь к найденному файлу WPP через общий механизм поиска"""
    return get_wpp_path()


def _normalize_header_cell(value: Any) -> str:
    """Приводит ячейку заголовка к строке без лишних пробелов"""
    return str(value).strip() if value is not None else ""


def _find_header_row(worksheet: Any, max_rows: int = 80) -> Tuple[int, List[str]]:
    """Находит строку заголовков на листе WPP по набору обязательных колонок

    В WPP2024 заголовок находится на фиксированной строке, но в следующих
    ревизиях ООН может изменить количество служебных строк. Поэтому строка
    определяется динамически, а не через жёстко заданный номер
    """
    max_scan = min(max_rows, worksheet.max_row or max_rows)
    for row_index in range(1, max_scan + 1):
        raw_values = next(worksheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        values = [_normalize_header_cell(value) for value in raw_values]
        if _REQUIRED_HEADER_NAMES.issubset(set(values)):
            return row_index, values
    raise KeyError(f"На листе '{worksheet.title}' не найдена строка заголовков WPP с нужными колонками.")


def _sheet_has_required_columns(worksheet: Any) -> bool:
    """Проверяет, похож ли лист Excel на сценарий WPP с демографическими показателями"""
    try:
        _find_header_row(worksheet)
        return True
    except Exception:
        return False


@lru_cache(maxsize=16)
def _get_available_forecast_sheets(path: str) -> List[str]:
    """Считывает из Excel-файла список доступных прогнозных сценариев"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        variants: List[str] = []
        for sheet_name in workbook.sheetnames:
            if sheet_name in _NON_FORECAST_SHEETS:
                continue
            worksheet = workbook[sheet_name]
            if _sheet_has_required_columns(worksheet):
                variants.append(sheet_name)
        if variants:
            return variants
        return [name for name in workbook.sheetnames if name not in _NON_FORECAST_SHEETS]
    finally:
        workbook.close()


def _to_numeric_people(series: pd.Series) -> pd.Series:
    """Переводит значения WPP из тысяч человек в обычные числовые значения"""
    cleaned = (
        series.replace("…", pd.NA)
        .astype(str)
        .str.replace(" ", "", regex=False)
        .str.replace(",", "", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce") * 1000


def _empty_un_forecast_frame() -> pd.DataFrame:
    """Создаёт пустую таблицу с теми же колонками, что и у подготовленного прогноза ООН"""
    return pd.DataFrame(
        columns=[
            "year",
            "variant",
            "country",
            "un_population_start",
            "un_births",
            "un_deaths",
            "un_net_migration",
            "un_inflow",
            "un_outflow",
            "un_total_forecast",
        ]
    )


def _cache_directory(path: str) -> str:
    """Возвращает каталог для локального кэша прочитанных строк WPP"""
    base_dir = os.path.dirname(os.path.abspath(path))
    cache_dir = os.path.join(base_dir, ".wpp_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def _safe_cache_part(value: str) -> str:
    """Преобразует название страны или сценария в безопасный фрагмент имени файла"""
    normalized = re.sub(r"[^A-Za-zА-Яа-я0-9_.-]+", "_", value.strip())
    return normalized[:80] or "value"


def _cache_path(path: str, variant: str, country: str) -> str:
    """Формирует путь к файлу кэша с учётом даты изменения исходного Excel-файла"""
    stat = os.stat(path)
    token = f"{os.path.abspath(path)}|{stat.st_mtime_ns}|{variant}|{country}"
    digest = hashlib.sha1(token.encode("utf-8")).hexdigest()[:16]
    filename = f"{_safe_cache_part(variant)}__{_safe_cache_part(country)}__{digest}.pkl"
    return os.path.join(_cache_directory(path), filename)


def _read_cached_country(path: str, variant: str, country: str) -> Optional[pd.DataFrame]:
    """Пытается загрузить уже подготовленные строки WPP из локального кэша"""
    filename = _cache_path(path, variant, country)
    if not os.path.exists(filename):
        return None
    try:
        with open(filename, "rb") as file:
            return pickle.load(file)
    except Exception:
        return None


def _write_cached_country(path: str, variant: str, country: str, df: pd.DataFrame) -> None:
    """Сохраняет подготовленные строки WPP в локальный кэш"""
    try:
        with open(_cache_path(path, variant, country), "wb") as file:
            pickle.dump(df, file, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        return


@lru_cache(maxsize=256)
def _read_un_forecast_country(path: str, variant: str, country: str) -> pd.DataFrame:
    """Считывает из WPP строки выбранной страны и сценария с кэшированием результата"""
    cached = _read_cached_country(path, variant, country)
    if cached is not None:
        return cached.copy()

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if variant not in workbook.sheetnames:
            available = ", ".join(_get_available_forecast_sheets(path)[:10])
            raise ValueError(f"В файле WPP нет листа '{variant}'. Доступные сценарии: {available}")
        worksheet = workbook[variant]

        header_row, header = _find_header_row(worksheet)
        indexes = {name: header.index(name) for name in _COLUMN_MAP.keys() if name in header}
        missing = set(_COLUMN_MAP.keys()) - set(indexes.keys())
        if missing:
            raise KeyError(f"В листе '{variant}' не найдены столбцы: {sorted(missing)}")

        rows = []
        found_country = False
        country_col = indexes["Region, subregion, country or area *"]
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_country = row[country_col]
            if row_country == country:
                found_country = True
                rows.append({new_name: row[indexes[old_name]] for old_name, new_name in _COLUMN_MAP.items()})
            elif found_country and row_country not in (None, country):
                break
    finally:
        workbook.close()

    df = pd.DataFrame(rows)
    if df.empty:
        _write_cached_country(path, variant, country, df)
        return df

    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    for col in _NUMERIC_COLUMNS:
        df[col] = _to_numeric_people(df[col])

    df = df.dropna(subset=["country", "year", "population"]).copy()
    _write_cached_country(path, variant, country, df)
    return df.copy()


def load_un_forecast(
    country: str,
    variant: str = "Medium variant",
    migration_policy: int = 2,
    start_year: int = 2024,
    end_year: Optional[int] = None,
) -> pd.DataFrame:
    """Загружает и подготавливает прогноз WPP для выбранной страны, сценария и периода"""
    path = get_wpp_path()
    variant = get_effective_un_forecast_variant(variant) or variant
    start_year = int(start_year)
    df = _read_un_forecast_country(path, variant, country).copy()
    if df.empty:
        return _empty_un_forecast_frame()

    df["year"] = df["year"].astype(int)
    df = df[df["year"] >= start_year]
    if end_year is not None:
        df = df[df["year"] <= int(end_year)]
    if df.empty:
        return _empty_un_forecast_frame()

    if migration_policy == 0:
        df["un_inflow"] = df["births"] + df["net_migration"]
        df["un_outflow"] = df["deaths"]
    elif migration_policy == 1:
        df["un_inflow"] = df["births"] + df["net_migration"].clip(lower=0)
        df["un_outflow"] = df["deaths"] + (-df["net_migration"]).clip(lower=0)
    else:
        df["un_inflow"] = df["births"]
        df["un_outflow"] = df["deaths"] - df["net_migration"]

    df["un_total_forecast"] = df["population"] + df["un_inflow"] - df["un_outflow"]

    return df.rename(
        columns={
            "population": "un_population_start",
            "births": "un_births",
            "deaths": "un_deaths",
            "net_migration": "un_net_migration",
        }
    )[
        [
            "year",
            "variant",
            "country",
            "un_population_start",
            "un_births",
            "un_deaths",
            "un_net_migration",
            "un_inflow",
            "un_outflow",
            "un_total_forecast",
        ]
    ].sort_values("year")
